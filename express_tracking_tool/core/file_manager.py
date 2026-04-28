"""
Excel文件管理器
"""

from typing import List, Tuple, Dict, Optional
import openpyxl
from openpyxl import Workbook
from pathlib import Path


class FileManager:
    """Excel文件管理器"""
    
    def __init__(self, file_path: str):
        """初始化文件管理器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.order_sheet = None
        self.courier_sheet = None
        self._tracking_number_col = None
        self._courier_col = None
        
    def validate_template(self) -> Tuple[bool, str]:
        """验证Excel模板格式
        
        Returns:
            (是否有效, 错误消息)
        """
        try:
            # 尝试打开文件
            if self.file_path.suffix.lower() == '.xlsx':
                self.workbook = openpyxl.load_workbook(self.file_path)
            else:
                return False, "不支持的文件格式，请使用.xlsx文件"
            
            # 检查工作表数量
            if len(self.workbook.worksheets) < 2:
                return False, "Excel文件必须包含至少2个工作表"
            
            # 获取工作表
            self.order_sheet = self.workbook.worksheets[0]
            self.courier_sheet = self.workbook.worksheets[1]
            
            # 验证第一个工作表的表头
            header_row = list(self.order_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            # 模糊匹配必需列
            required_keywords = ["订单编号", "物流公司", "运单号"]
            found_columns = {keyword: None for keyword in required_keywords}
            
            for col_idx, cell_value in enumerate(header_row, 1):
                if cell_value:
                    cell_str = str(cell_value)
                    for keyword in required_keywords:
                        if keyword in cell_str and found_columns[keyword] is None:
                            found_columns[keyword] = col_idx
            
            # 检查是否所有必需列都找到
            missing_columns = [k for k, v in found_columns.items() if v is None]
            if missing_columns:
                return False, f"缺少必需列: {', '.join(missing_columns)}"
            
            # 保存列索引
            self._tracking_number_col = found_columns["运单号"]
            self._courier_col = found_columns["物流公司"]
            
            # 检查第二个工作表
            if self.courier_sheet.max_row < 2:  # 至少有表头和一行数据
                return False, "第二个工作表必须包含快递公司列表"
            
            return True, ""
            
        except Exception as e:
            return False, f"文件验证失败: {str(e)}"
    
    def read_tracking_numbers(self) -> List[Tuple[int, str]]:
        """读取运单号数据
        
        Returns:
            [(行号, 运单号), ...]
        """
        if not self.order_sheet or not self._tracking_number_col:
            return []
        
        tracking_numbers = []
        
        # 从第2行开始读取数据（跳过表头）
        for row_idx in range(2, self.order_sheet.max_row + 1):
            cell = self.order_sheet.cell(row=row_idx, column=self._tracking_number_col)
            if cell.value:
                tracking_number = str(cell.value).strip()
                if tracking_number:
                    tracking_numbers.append((row_idx, tracking_number))
        
        return tracking_numbers
    
    def read_courier_list(self) -> list:
        """读取快递公司列表（名称和编码）
        
        Returns:
            [(快递公司名称, 编码), ...]
        """
        if not self.courier_sheet:
            return []
        
        courier_list = []
        
        # 从第2行开始读取（跳过表头"快递名称"/"代码"）
        for row_idx in range(2, self.courier_sheet.max_row + 1):
            name_cell = self.courier_sheet.cell(row=row_idx, column=1)
            code_cell = self.courier_sheet.cell(row=row_idx, column=2)
            if name_cell.value and code_cell.value is not None:
                name = str(name_cell.value).strip()
                code = str(code_cell.value).strip()
                if name and code:
                    courier_list.append((name, code))
        
        return courier_list
    
    def write_courier_codes(self, results: Dict[int, str]) -> None:
        """写入快递公司编码
        
        Args:
            results: {行号: 快递公司编码}
        """
        if not self.order_sheet or not self._courier_col:
            return
        
        for row_idx, courier_code in results.items():
            cell = self.order_sheet.cell(row=row_idx, column=self._courier_col)
            cell.value = courier_code
    
    def save(self) -> None:
        """保存Excel文件"""
        if self.workbook:
            self.workbook.save(self.file_path)
    
    def close(self) -> None:
        """关闭Excel文件"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None